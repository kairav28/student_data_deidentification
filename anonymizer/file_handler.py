"""In-memory file reading and writing for CSV and Excel (.xlsx) files.

No uploaded or processed data is ever written to disk. All work happens
against in-memory bytes buffers (BytesIO) and pandas DataFrames.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook

from .validator import (
    MESSAGES,
    EmptyDatasetError,
    EmptyFileError,
    FileReadError,
    MultiSheetExcelError,
    NoColumnsError,
    UnsupportedFileTypeError,
)

ALLOWED_EXTENSIONS = (".csv", ".xlsx")


def get_extension(filename: str) -> str:
    lower = filename.lower()
    for ext in ALLOWED_EXTENSIONS:
        if lower.endswith(ext):
            return ext
    return ""


def load_dataset(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Load an uploaded CSV or single-sheet Excel file into a DataFrame.

    Raises a DeidentificationError subclass with a user-facing message on
    failure. The file is never written to disk during this process.
    """
    extension = get_extension(filename)
    if extension not in ALLOWED_EXTENSIONS:
        raise UnsupportedFileTypeError(MESSAGES[UnsupportedFileTypeError])

    if len(file_bytes) == 0:
        raise EmptyFileError(MESSAGES[EmptyFileError])

    if extension == ".csv":
        df = _load_csv(file_bytes)
    else:
        df = _load_xlsx(file_bytes)

    if df.shape[1] == 0:
        raise NoColumnsError(MESSAGES[NoColumnsError])
    if df.shape[0] == 0:
        raise EmptyDatasetError(MESSAGES[EmptyDatasetError])

    return df


def _load_csv(file_bytes: bytes) -> pd.DataFrame:
    try:
        return pd.read_csv(io.BytesIO(file_bytes))
    except pd.errors.EmptyDataError as exc:
        raise EmptyFileError(MESSAGES[EmptyFileError]) from exc
    except Exception as exc:
        raise FileReadError(MESSAGES[FileReadError]) from exc


def _load_xlsx(file_bytes: bytes) -> pd.DataFrame:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames)
        workbook.close()
    except Exception as exc:
        raise FileReadError(MESSAGES[FileReadError]) from exc

    if len(sheet_names) != 1:
        raise MultiSheetExcelError(MESSAGES[MultiSheetExcelError])

    try:
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_names[0], engine="openpyxl")
    except Exception as exc:
        raise FileReadError(MESSAGES[FileReadError]) from exc


def write_dataset(df: pd.DataFrame, extension: str) -> tuple[bytes, str]:
    """Serialise a DataFrame to bytes matching the given extension.

    Returns (file_bytes, mime_type). Nothing is written to disk.
    """
    buffer = io.BytesIO()
    if extension == ".csv":
        df.to_csv(buffer, index=False)
        mime_type = "text/csv"
    elif extension == ".xlsx":
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        raise UnsupportedFileTypeError(MESSAGES[UnsupportedFileTypeError])

    buffer.seek(0)
    return buffer.getvalue(), mime_type


def output_filename(original_filename: str, extension: str) -> str:
    stem = original_filename
    for ext in ALLOWED_EXTENSIONS:
        if stem.lower().endswith(ext):
            stem = stem[: -len(ext)]
            break
    return f"{stem}_deidentified{extension}"
