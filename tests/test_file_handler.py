import io

import pandas as pd
import pytest
from openpyxl import Workbook

from anonymizer.file_handler import get_extension, load_dataset, output_filename, write_dataset
from anonymizer.validator import (
    EmptyDatasetError,
    EmptyFileError,
    MultiSheetExcelError,
    NoColumnsError,
    UnsupportedFileTypeError,
)


def _csv_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    df.to_csv(buffer, index=False)
    return buffer.getvalue()


def _xlsx_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    return buffer.getvalue()


def test_load_csv_input():
    df = pd.DataFrame({"Name": ["Rahul", "Aisha"], "Score": [1, 2]})
    loaded = load_dataset(_csv_bytes(df), "students.csv")
    assert list(loaded.columns) == ["Name", "Score"]
    assert len(loaded) == 2


def test_load_single_sheet_xlsx_input():
    df = pd.DataFrame({"Name": ["Rahul", "Aisha"], "Score": [1, 2]})
    loaded = load_dataset(_xlsx_bytes(df), "students.xlsx")
    assert list(loaded.columns) == ["Name", "Score"]
    assert len(loaded) == 2


def test_multi_sheet_xlsx_is_rejected():
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.create_sheet("Sheet2")
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(MultiSheetExcelError):
        load_dataset(buffer.getvalue(), "students.xlsx")


def test_unsupported_file_type_is_rejected():
    with pytest.raises(UnsupportedFileTypeError):
        load_dataset(b"some content", "students.txt")


def test_empty_file_is_rejected():
    with pytest.raises(EmptyFileError):
        load_dataset(b"", "students.csv")


def test_empty_dataset_is_rejected():
    df = pd.DataFrame({"Name": [], "Score": []})
    with pytest.raises(EmptyDatasetError):
        load_dataset(_csv_bytes(df), "students.csv")


def test_no_columns_is_rejected():
    workbook = Workbook()  # single, completely blank default sheet
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(NoColumnsError):
        load_dataset(buffer.getvalue(), "students.xlsx")


def test_output_format_matches_csv_input():
    df = pd.DataFrame({"Name": ["ANON_00001"], "Score": [10]})
    output_bytes, mime = write_dataset(df, ".csv")
    assert mime == "text/csv"
    result = pd.read_csv(io.BytesIO(output_bytes))
    assert list(result.columns) == ["Name", "Score"]


def test_output_format_matches_xlsx_input():
    df = pd.DataFrame({"Name": ["ANON_00001"], "Score": [10]})
    output_bytes, mime = write_dataset(df, ".xlsx")
    assert mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    result = pd.read_excel(io.BytesIO(output_bytes), engine="openpyxl")
    assert list(result.columns) == ["Name", "Score"]


def test_get_extension_is_case_insensitive():
    assert get_extension("Students.CSV") == ".csv"
    assert get_extension("Students.XLSX") == ".xlsx"
    assert get_extension("students.txt") == ""


def test_output_filename_preserves_extension():
    assert output_filename("students.csv", ".csv") == "students_deidentified.csv"
    assert output_filename("students.xlsx", ".xlsx") == "students_deidentified.xlsx"
